import openpyxl
wb = openpyxl.load_workbook('updates.xlsx')
for sn in wb.sheetnames:
    print('SHEET:', sn)
    sheet = wb[sn]
    for row in sheet.iter_rows(values_only=True, max_row=10):
        print(row)
